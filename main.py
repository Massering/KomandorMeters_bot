import os
from datetime import datetime as dt, timedelta
from time import sleep

import json
from openpyxl import styles, Workbook
import sqlite3
import telebot
from telebot import types

# Имена файлов, использующихся в программе
DB_NAME = "Table.sqlite"
EXCEL_TABLE = "temp_table.xlsx"
USERS_NAME = 'users.json'
COMPANIES_NAME = 'companies.json'

# Константы, по совместительству заголовки в таблице Excel
# Лучше не менять, иначе файл users.json нужно будет перезаписать
HEADERS = ['Company', 'Address', 'Username', 'User_id', 'Phone', 'Counter', 'Data', 'Datetime']
COMPANY, ADDRESS, USERNAME, USER_ID, PHONE, COUNTER, DATA, DATETIME = HEADERS

POSITIVE_ANSWERS = ['yes', 'y', 'да', 'д', '1', 'дп', 'lf']  # Ответы, которые мы принимаем за положительный ответ
RUS = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'
ENG = 'abcdefghijklmnopqrstuvwxyz'
DIGITS = '1234567890'
ALLOWED_SIMBOLS = DIGITS + ENG + ENG.upper() + RUS + RUS.upper() + ' -.,()/+_'


def dump(obj, filename):
    """Функция для простого внесения данных в файлы"""
    json.dump(obj, open(filename, 'w', encoding='UTF-8'), ensure_ascii=False, indent=4)


#          Sophia,     Maksim
ADMINS = [979923466, 1089524173]
TOKEN = '5253767532:AAF9DZ2obpuMVKiQD_VHmskA5WTtkjkys3k'

bot = telebot.TeleBot(TOKEN)

# Открытие БД
conn = sqlite3.connect(DB_NAME, check_same_thread=False)

# Создание и открытие json файла со списком зарегистрированных пользователей
if not os.path.exists(USERS_NAME):
    dump({}, USERS_NAME)
users = json.load(open(USERS_NAME, 'r', encoding='utf-8'))

# Создание и открытие файла со списком компаний
if not os.path.exists(COMPANIES_NAME):
    dump({}, COMPANIES_NAME)
companies = json.load(open(COMPANIES_NAME, 'r', encoding='UTF-8'))

# Словарь, в котором будет содержаться информация, которую вносят пользователи в данный момент
recording_data = {}


def get_date() -> str:
    """Получение строки с датой и временем сейчас"""
    return dt.now().strftime("%Y-%m-%d %H:%M:%S")


def get_changes(old, new) -> str:
    """Получение строки с изменениями в словаре. Проверяет только значения, существующие в обоих словарях"""
    data = [(key, old[key], new[key]) for key in old if key in new]
    return "\n".join([f"{i[0]}: {i[1] + ' => ' + i[2] if i[1] != i[2] else i[1]}" for i in data])


def check_number(number) -> str:
    """Изменение номера телефона по формату"""
    number = ''.join(number.replace('(', '').replace(')', '').replace('-', '').split())
    if number[0] == '8':
        number = '+7' + number[1:]
    assert len(number) == 12
    assert number[:2] == '+7'
    assert number[2:].isdigit()
    return number


def check_data(data) -> bool:
    """Проверка, соответствуют ли введенные данные ПУ допустимому формату"""
    return all(i in '1234567890. ' for i in data)


def log(message, symbols=ALLOWED_SIMBOLS, start_call=False) -> bool:
    """Вывод в консоль уведомления о сообщении боту + Проверка сообщения (выход, атака)"""
    name = message.from_user.username or message.from_user.first_name or 'Unknown'
    if str(message.from_user.id) not in users:
        name += f' (id {message.from_user.id})'
    print(f'{get_date()} - {name}: "{message.text}"')
    # print(recording_data)

    if not message.text:
        error_text = 'Текст сообщения не должен быть пустым'
    elif message.text == '/exit':
        error_text = 'Выход в меню'
    elif '/' in message.text and not start_call:
        start(message)
        return True

    elif any(i not in symbols for i in message.text):
        error_text = 'Сообщение содержит недопустимые символы'
    elif len(message.text) > 255:
        error_text = 'Сообщение не должно превышать длину 255 символов'
    else:
        return False

    bot.send_message(message.from_user.id, error_text)
    print_commands(message)
    return True


def make_bool_keyboard(one_time=True):
    """Возвращает клавиатуру, состоящую из кнопок "Да" и "Нет\""""
    keyboard = types.ReplyKeyboardMarkup(True, one_time)
    keyboard.add(types.KeyboardButton('Да'), types.KeyboardButton('Нет'))
    return keyboard


def make_keyboard(values, one_time=True):
    """Возвращает клавиатуру, содержащую кнопки со значениями values"""
    keyboard = types.ReplyKeyboardMarkup(True, one_time)
    for value in values:
        key1 = types.KeyboardButton(value)
        keyboard.add(key1)
    return keyboard


def print_commands(message):
    # TODO: /get_users - просмотр всех зарегистрированных пользователей
    text = f'''Воспользуйтесь функциями меню:
/createentry - внесение показания прибора учета
/get_entries - получение записанных показаний по приборам учета
/add_counter - регистрация прибора учета по вашему адресу
/edit_user - редактирование вашего профиля
/exit - завершения работы'''
    if message.from_user.id in ADMINS:
        text += '''\n
/get_records - получение показаний за месяц в Excel таблице
/add_company - добавление компании
/get_companies - просмотр всех зарегистрированных компаний
/remove_user - удаление зарегистрированного пользователя по id
/edit_user_by_id - редактирование данных зарегистрированного пользователя по id'''
    bot.send_message(message.from_user.id, text)


@bot.message_handler(content_types=['text'])
def start(message):
    """Изначальная функция, принимающая запросы пользователя"""
    if log(message, start_call=True):
        return

    user_id = message.from_user.id

    if str(user_id) not in users:
        if message.text == '/edit_user':
            message.text = 'Да'
            if_registration(message)
            return
        else:
            bot.send_message(user_id, "Вы не зарегистрированы. Хотите зарегистрироваться?",
                             reply_markup=make_bool_keyboard())
            bot.register_next_step_handler(message, if_registration)

    elif message.text == '/createentry':
        create_entry(message)

    elif message.text == '/edit_user':
        # Подтверждаем регистрацию
        bot.send_message(user_id, 'Вы уже зарегистрированы. Ваши данные будут заменены. Вы уверены?',
                         reply_markup=make_bool_keyboard())
        bot.register_next_step_handler(message, if_registration)

    elif message.text == '/get_entries':
        get_entries(message)

    elif message.text == '/add_counter':
        bot.send_message(user_id, 'Введите номер регистрируемого прибора учёта')
        bot.register_next_step_handler(message, add_counter)

    elif user_id in ADMINS and message.text == '/add_company':
        bot.send_message(user_id, 'Введите название регистрируемой компании')
        bot.register_next_step_handler(message, add_company)

    elif user_id in ADMINS and message.text == '/get_companies':
        get_companies(message)

    elif user_id in ADMINS and message.text == '/remove_user':
        bot.send_message(user_id, 'Введите id пользователя')
        bot.register_next_step_handler(message, remove_user_by_id)

    elif user_id in ADMINS and message.text == '/edit_user_by_id':
        bot.send_message(user_id, 'Введите id пользователя')
        bot.register_next_step_handler(message, edit_user_by_id)

    elif user_id in ADMINS and message.text == '/get_records':
        get_records(message)

    # Обработка сообщений, не содержащих команд
    else:
        print_commands(message)


def create_entry(message):
    user_id = message.from_user.id
    user = users[str(user_id)]

    # Предлагаем пользователю список счётчиков по этому адресу
    counters = companies[user[COMPANY]][user[ADDRESS]]

    # Если счетчик всего один, не спрашиваем пользователя
    if len(counters) == 0:
        bot.send_message(user_id, 'Нет зарегистрированных приборов учёта. '
                                  'Для регистрации введите /add_counter')
        return

    if len(counters) == 1:
        message.text = list(counters)[0]
        get_counter(message)
        return

    bot.send_message(user_id, "Выберите номер прибора учёта из списка",
                     reply_markup=make_keyboard(counters))
    bot.register_next_step_handler(message, get_counter)


def get_records(message):
    date_from = dt.now() - timedelta(days=30)
    date_to = dt.now()
    cursor = conn.cursor()
    request = (f"SELECT * FROM records WHERE {DATETIME} BETWEEN '{date_from.strftime('%Y-%m-%d')}'"
               f" AND '{get_date()}' ORDER BY {DATETIME}")
    result = cursor.execute(request).fetchall()
    cursor.close()

    workbook = Workbook()
    sheet = workbook.worksheets[0]
    for j, header in enumerate(HEADERS, 1):
        cell = sheet.cell(1, j)
        cell.value = header.capitalize()
        cell.font = styles.Font(bold=True)
        cell.alignment = styles.Alignment(horizontal='center')

    for i, record in enumerate(result, 2):
        for j, value in enumerate(record[1:], 1):
            cell = sheet.cell(i, j)
            cell.value = value

    workbook.save(EXCEL_TABLE)
    workbook.close()

    filename = f"Records for {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}"
    bot.send_document(message.chat.id, open(EXCEL_TABLE, 'rb').read(),
                      visible_file_name=filename + '.xlsx')
    os.remove(EXCEL_TABLE)


def get_companies(message):
    """Администратору выводится список всех компаний"""
    if companies:
        text = 'Список всех зарегистрированных компаний:\n'
        text += '\n'.join(companies)
    else:
        text = 'Нет зарегистрированных компаний'
    bot.send_message(message.from_user.id, text)


def get_entries(message):
    """Пользователю выводится список всех счётчиков и их текущих показаний"""
    user = users[str(message.from_user.id)]
    if companies[user[COMPANY]][user[ADDRESS]]:
        text = 'Внесенные показания по вашим приборам учета:\n'
        for i in companies[user[COMPANY]][user[ADDRESS]].items():
            text += f'"{i[0]}": {i[1] or "Нет показаний"}\n'
    else:
        text = 'Нет зарегистрированных приборов учета'
    bot.send_message(message.from_user.id, text)


def edit_user_by_id(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        return

    if message.text == 'self':
        message.text = str(message.from_user.id)

    if message.text not in users:
        bot.send_message(message.from_user.id, 'Пользователь с этим id не зарегистрирован')
        print_commands(message)
        return

    recording_data[message.from_user.id] = {USER_ID: message.text}

    data = "\n".join([f"{i[0]}: {i[1]}" for i in users[message.text].items()])
    bot.send_message(message.from_user.id, f'Текущие данные пользователя:\n{data}')

    companies_list = [users[message.text][COMPANY]] + [i for i in companies if i != users[message.text][COMPANY]]
    bot.send_message(message.from_user.id, f'Введите компанию, к которой должен быть привязан пользователь',
                     reply_markup=make_keyboard(companies_list))
    bot.register_next_step_handler(message, edit_user_by_id_company)


def edit_user_by_id_company(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text not in companies:
        bot.send_message(message.from_user.id, 'Компания с таким названием не зарегистрирована. Попробуйте ещё')
        bot.register_next_step_handler(message, edit_user_by_id_company)
        return

    recording_data[message.from_user.id][COMPANY] = message.text

    user = users[recording_data[message.from_user.id][USER_ID]]

    if message.text == user[COMPANY]:
        addresses = [user[ADDRESS]] + [i for i in companies[user[COMPANY]] if i != user[ADDRESS]]
    else:
        addresses = companies[message.text]
    bot.send_message(message.from_user.id, f'Введите адрес, где работает пользователь',
                     reply_markup=make_keyboard(addresses))
    bot.register_next_step_handler(message, edit_user_by_id_address)


def edit_user_by_id_address(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    cur_data = recording_data[message.from_user.id]

    if message.text not in companies[cur_data[COMPANY]]:
        bot.send_message(message.from_user.id, 'Этот адрес будет внесен в список')

    cur_data[ADDRESS] = message.text

    bot.send_message(message.from_user.id, f'Введите телефон пользователя',
                     reply_markup=make_keyboard([users[cur_data[USER_ID]][PHONE]]))
    bot.register_next_step_handler(message, edit_user_by_id_phone)


def edit_user_by_id_phone(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    cur_data = recording_data[message.from_user.id]

    try:
        cur_data[PHONE] = check_number(message.text)
    except (AssertionError, TypeError, ValueError):
        bot.send_message(message.from_user.id, 'Телефон не соответствует формату. Попробуйте ещё раз')
        bot.register_next_step_handler(message, edit_user_by_id_phone)
        return

    bot.send_message(message.from_user.id, f'Введите имя пользователя',
                     reply_markup=make_keyboard([users[cur_data[USER_ID]][USERNAME]]))
    bot.register_next_step_handler(message, edit_user_by_id_username)


def edit_user_by_id_username(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    cur_data = recording_data[message.from_user.id]

    cur_data[USERNAME] = message.text

    user = users[cur_data[USER_ID]]
    changes = get_changes(user, cur_data)
    bot.send_message(message.from_user.id, f'Вы подтверждаете изменения?\n{changes}',
                     reply_markup=make_bool_keyboard())

    bot.register_next_step_handler(message, edit_user_by_id_verification)


def edit_user_by_id_verification(message):
    """Подтверждение внесения данных у пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id
    cur_data = recording_data[user_id]

    if message.text.lower() in POSITIVE_ANSWERS:
        if cur_data[ADDRESS] not in companies[cur_data[COMPANY]]:
            companies[cur_data[COMPANY]][cur_data[ADDRESS]] = {}
            dump(companies, COMPANIES_NAME)

        cur_user_id = cur_data[USER_ID]
        changes = get_changes(users[cur_user_id], cur_data)

        # Перезапись файла
        users[cur_user_id] = cur_data.copy()
        dump(users, USERS_NAME)
        del recording_data[message.from_user.id]

        bot.send_message(user_id, 'Вы успешно изменили данные пользователя')
        print_commands(message)
        text = f'Данные пользователя id{cur_user_id} были изменены администратором id{user_id}:\n{changes}'

        # Отправляем информацию об изменениях
        for admin_id in set(ADMINS + [int(cur_user_id)]):
            if admin_id != user_id:
                bot.send_message(admin_id, text)
    else:
        del recording_data[message.from_user.id]
        bot.send_message(user_id, 'Возврат в меню')
        print_commands(message)


def remove_user_by_id(message):
    """Администратор удаляет зарегистрированного пользователя по его id"""
    if log(message):
        return

    if message.text == 'self':
        message.text = str(message.from_user.id)

    if message.text not in users:
        bot.send_message(message.from_user.id, 'Пользователь с этим id не зарегистрирован')
        print_commands(message)
        return

    recording_data[message.from_user.id] = message.text

    data = "\n".join([": ".join(map(str, i)) for i in [('id', message.text)] + list(users[message.text].items())])
    bot.send_message(message.from_user.id, f'Вы действительно хотите удалить пользователя с данными:\n{data}',
                     reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, remove_user_by_id_verification)


def remove_user_by_id_verification(message):
    """Подтверждение удаления пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        del_user_id = recording_data[message.from_user.id]

        del users[del_user_id]
        dump(users, USERS_NAME)

        bot.send_message(int(del_user_id), 'Вы были удалены администратором.')

        for admin_id in ADMINS:
            if admin_id == message.from_user.id:
                text = f'Пользователь {del_user_id} был удален.'
            else:
                text = f'Пользователь {del_user_id} был удален администратором id{message.from_user.id}.'
            bot.send_message(admin_id, text)

        del recording_data[message.from_user.id]

    else:
        del recording_data[message.from_user.id]
        bot.send_message(message.from_user.id, 'Хорошо')


def add_company(message):
    """Регистрация компании администратором"""
    if log(message):
        return

    recording_data[message.from_user.id] = message.text
    bot.send_message(message.from_user.id, f'Зарегистрировать компанию "{message.text}"?',
                     reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, add_company_verification)


def add_company_verification(message):
    """Подтверждение регистрации компании"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        companies[recording_data[message.from_user.id]] = {}
        dump(companies, COMPANIES_NAME)

        del recording_data[message.from_user.id]

        bot.send_message(message.from_user.id, 'Компания зарегистрирована')

    else:
        bot.send_message(message.from_user.id, 'Компания не зарегистрирована')
        message.text = 'exit_code_1'
        start(message)


def add_counter(message):
    """Регистрация прибора учета пользователем"""
    if log(message):
        return

    recording_data[message.from_user.id] = message.text
    bot.send_message(message.from_user.id, f'Зарегистрировать прибор учета с номером "{message.text}" '
                                           f'по адресу "{users[str(message.from_user.id)][ADDRESS]}"?',
                     reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, add_counter_verification)


def add_counter_verification(message):
    """Подтверждение регистрации прибора учета пользователем"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        user = users[str(message.from_user.id)]
        companies[user[COMPANY]][user[ADDRESS]][recording_data[message.from_user.id]] = ""
        dump(companies, COMPANIES_NAME)

        del recording_data[message.from_user.id]

        bot.send_message(message.from_user.id, 'Прибор учета зарегистрирован')

    else:
        bot.send_message(message.from_user.id, 'Прибор учета не зарегистрирован')
    print_commands(message)


def get_counter(message):
    """Получение названия/номера счётчика у пользователя"""
    if log(message):
        return

    counter = message.text
    user = users[str(message.from_user.id)]

    if counter in companies[user[COMPANY]][user[ADDRESS]]:
        value = companies[user[COMPANY]][user[ADDRESS]][counter]
        if value:
            bot.send_message(message.from_user.id, f'Прошлое показание прибора учёта "{counter}": {value}')
        else:
            bot.send_message(message.from_user.id, f'Нет предыдущих показаний по счётчику "{counter}"')

    else:
        bot.send_message(message.from_user.id, 'Прибор учёта с таким номером не зарегистрирован. '
                                               'Для регистрации введите /add_counter')
        return

    cur_data = recording_data[message.from_user.id] = {}
    cur_data[COMPANY] = user[COMPANY]
    cur_data[ADDRESS] = user[ADDRESS]
    cur_data[USERNAME] = user[USERNAME]
    cur_data[USER_ID] = str(message.from_user.id)
    cur_data[PHONE] = user[PHONE]
    cur_data[COUNTER] = message.text

    bot.send_message(message.from_user.id, f'Введите текущее показание прибора учёта с номером "{message.text}"')
    bot.register_next_step_handler(message, get_data)


def get_data(message):
    """Получение данных счётчика у пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if check_data(message.text):
        cur_data = recording_data[message.from_user.id]
        cur_data[DATA] = message.text
        cur_data[DATETIME] = get_date()

        s = "\n".join([": ".join(map(str, i)) for i in cur_data.items()])
        bot.send_message(message.from_user.id, f'Полученные данные: \n{s}')
        bot.send_message(message.from_user.id, f'Всё верно?', reply_markup=make_bool_keyboard())
        bot.register_next_step_handler(message, data_verification)

    else:
        bot.send_message(message.from_user.id, f'Введенные данные должны быть целым числом. Введите ещё раз')
        bot.register_next_step_handler(message, get_data)


def data_verification(message):
    """Подтверждение, что пользователь хочет внести данные"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() not in POSITIVE_ANSWERS:
        del recording_data[message.from_user.id]
        bot.send_message(message.from_user.id, f'Данные не записаны')
        return

    cur_data = recording_data[message.from_user.id]

    companies[cur_data[COMPANY]][cur_data[ADDRESS]][cur_data[COUNTER]] = cur_data[DATA]
    dump(companies, COMPANIES_NAME)

    cursor = conn.cursor()
    # COMPANY, ADDRESS, USERNAME, USER_ID, PHONE, COUNTER, DATA, DATETIME = HEADERS
    cursor.execute(f"INSERT INTO records ({', '.join(HEADERS)}) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                   [cur_data[header] for header in HEADERS])
    cursor.close()
    conn.commit()

    del recording_data[message.from_user.id]

    bot.send_message(message.from_user.id, f'Данные записаны')
    print_commands(message)


def if_registration(message):
    """Проверка, точно ли пользователь хочет зарегистрироваться
    Регистрация происходит в несколько этапов.
    Компания => Номер телефона => Имя
    Эти данные впоследствии будут заноситься в таблицу, когда этот пользователь вводит показания"""
    if log(message):
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        recording_data[message.from_user.id] = {}

        if str(message.from_user.id) in users:
            markup = make_keyboard([users[str(message.from_user.id)][COMPANY]])
        else:
            markup = None

        bot.send_message(message.from_user.id, 'Введите название своей компании',
                         reply_markup=markup)
        bot.register_next_step_handler(message, register_company)

    else:
        bot.send_message(message.from_user.id, 'Ок. Не хотите - не надо')


def register_company(message):
    """Получение компании, в которой работает пользователь"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text not in companies:
        bot.send_message(message.from_user.id, 'Такой компании нет. Введите ещё раз')
        bot.register_next_step_handler(message, register_company)
        return

    recording_data[message.from_user.id][COMPANY] = message.text

    addresses = companies[message.text]
    bot.send_message(message.from_user.id, 'Выберите адрес своего магазина из списка или, '
                                           'если его нет, введите вручную',
                     reply_markup=make_keyboard(addresses))
    bot.register_next_step_handler(message, register_address)


def register_address(message):
    """Получение адреса, по которому работает пользователь"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    recording_data[message.from_user.id][ADDRESS] = message.text

    if str(message.from_user.id) in users:
        markup = make_keyboard([users[str(message.from_user.id)][PHONE]], False)
    else:
        markup = None

    bot.send_message(message.from_user.id, 'Введите свой номер телефона в федеральном формате (+7**********)',
                     reply_markup=markup)
    bot.register_next_step_handler(message, register_phone)


def register_phone(message):
    """Получение номера телефона пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    try:
        recording_data[message.from_user.id][PHONE] = check_number(message.text)

        # Предлагаем ему имя, под которым он зарегистрирован в Телеграмме
        # и также имя, под которым он был зарегистрирован в прошлый раз
        names = [((message.from_user.last_name or ' ') + ' ' + (message.from_user.first_name or ' ')).strip()]
        if str(message.from_user.id) in users and users[str(message.from_user.id)][USERNAME] != names[0]:
            names.insert(0, users[str(message.from_user.id)][USERNAME])

        bot.send_message(message.from_user.id, 'Введите своё имя', reply_markup=make_keyboard(names, False))
        bot.register_next_step_handler(message, register_name)

    except (AssertionError, IndexError, ValueError):
        bot.send_message(message.from_user.id, 'Номер введён неправильно. Введите ещё раз')
        bot.register_next_step_handler(message, register_phone)


def register_name(message):
    """Получение имени пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    recording_data[message.from_user.id][USERNAME] = message.text

    changes = get_changes(users[str(message.from_user.id)], recording_data[message.from_user.id])
    bot.send_message(message.from_user.id, f'Изменения:\n{changes}')
    bot.send_message(message.from_user.id, f'Внести их?', reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, register_verification)


def register_verification(message):
    """Подтверждение внесения данных у пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id
    cur_data = recording_data[user_id]

    if message.text.lower() in POSITIVE_ANSWERS:
        if cur_data[ADDRESS] not in companies[cur_data[COMPANY]]:
            companies[cur_data[COMPANY]][cur_data[ADDRESS]] = {}
            dump(companies, COMPANIES_NAME)

        if str(user_id) in users:
            changes = get_changes(users[str(user_id)], cur_data)

            bot.send_message(user_id, 'Вы успешно изменили данные')
            text = f'Пользователь id{user_id} изменил свои данные:\n{changes}'

        else:
            data = "\n".join([f"{i[0]}: {i[1]}" for i in cur_data.items()])
            bot.send_message(user_id, 'Вы успешно зарегистрировались')
            text = f'Пользователь id{user_id} зарегистрировался:\n{data}'
        print_commands(message)

        # Перезапись файла
        users[str(user_id)] = cur_data.copy()
        dump(users, USERS_NAME)
        del recording_data[message.from_user.id]

        # Отправляем админу информацию о регистрации
        for admin_id in ADMINS:
            bot.send_message(admin_id, text)
    else:
        del recording_data[message.from_user.id]
        bot.send_message(user_id, 'Возврат в меню')
        print_commands(message)


if __name__ == "__main__":
    while 1:
        try:
            bot.polling(none_stop=True, interval=0)
        except Exception as error:
            print(f'{get_date()} - FATAL_ERROR({error.__class__}, {error.__cause__}): {error}')
            sleep(1)
